"""Intake & Parser agent.

Accepts either a path to a `.xlsx`/`.txt` file or a raw string of questions
(one per line) and emits structured `Question` objects with a heuristic
category classification.
"""

from __future__ import annotations

import re
import uuid
from pathlib import Path
from typing import Iterable, List

from models import Question, QuestionCategory

_CATEGORY_KEYWORDS: dict[QuestionCategory, tuple[str, ...]] = {
    "certification": (
        "soc 2", "soc2", "iso 27001", "iso27001", "hipaa", "pci", "fedramp",
        "gdpr certified", "certif", "attestation", "audit report",
    ),
    "data-privacy": (
        "data stored", "data storage", "where is", "data residency",
        "subprocessor", "sub-processor", "retention", "delete",
        "personal data", "pii", "phi",
    ),
    "legal": (
        "guarantee", "warrant", "liability", "indemnif", "contract",
        "terms", "dpa ", "data processing agreement", "msa", "sla credit",
        "lawsuit",
    ),
    "technical": (
        "encrypt", "tls", "ssl", "mfa", "sso", "saml", "oauth", "vpc",
        "firewall", "waf", "key management", "kms", "backup", "rpo", "rto",
        "vulnerability", "pen test", "logging", "monitoring", "uptime",
        "availability",
    ),
}


def _classify(text: str) -> QuestionCategory:
    lower = text.lower()
    # Certification is highest precedence — bare mentions of HIPAA/SOC2 should
    # route there even if other keywords appear.
    for cat in ("certification", "legal", "data-privacy", "technical"):
        for kw in _CATEGORY_KEYWORDS[cat]:  # type: ignore[index]
            if kw in lower:
                return cat  # type: ignore[return-value]
    return "general"


def _new_id() -> str:
    return f"Q-{uuid.uuid4().hex[:8]}"


def _iter_lines(raw_input: str) -> Iterable[str]:
    """Split raw text into candidate question strings.

    Splits on newlines and also on terminal sentence punctuation in single-line
    inputs so paste-from-doc workflows still parse cleanly.
    """
    if "\n" in raw_input:
        for line in raw_input.splitlines():
            yield line
        return
    # Single-line input — split on sentence terminators while keeping question marks.
    for part in re.split(r"(?<=[?.!])\s+", raw_input):
        yield part


def _clean(line: str) -> str:
    line = line.strip()
    line = re.sub(r"^\d+[\.\)]\s*", "", line)  # strip "1." / "1)" prefixes
    line = re.sub(r"^[-*]\s*", "", line)  # strip bullet markers
    return line.strip()


def parse_text(raw_input: str) -> List[Question]:
    """Parse a raw multi-line string into structured Questions."""
    questions: List[Question] = []
    for line in _iter_lines(raw_input):
        cleaned = _clean(line)
        if not cleaned:
            continue
        # Skip headers / section labels that don't look like questions or statements.
        if len(cleaned) < 4:
            continue
        questions.append(
            Question(id=_new_id(), text=cleaned, category=_classify(cleaned))
        )
    return questions


def _parse_xlsx(path: Path) -> List[Question]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    questions: List[Question] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if not isinstance(cell, str):
                    continue
                cleaned = _clean(cell)
                if not cleaned or len(cleaned) < 4:
                    continue
                # Heuristic: skip obvious headers like "Question" / "Answer".
                if cleaned.lower() in {"question", "questions", "answer", "response"}:
                    continue
                questions.append(
                    Question(id=_new_id(), text=cleaned, category=_classify(cleaned))
                )
    return questions


def parse_questionnaire(source: str | Path) -> List[Question]:
    """Parse a `.xlsx`/`.txt` file path or a raw string into Questions."""
    if isinstance(source, Path) or (isinstance(source, str) and Path(source).is_file()):
        path = Path(source)
        if path.suffix.lower() == ".xlsx":
            return _parse_xlsx(path)
        return parse_text(path.read_text(encoding="utf-8"))
    return parse_text(source)
