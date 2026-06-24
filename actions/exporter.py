"""Final-action exporter: builds the filled questionnaire workbook and run summary."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import COMPANY_NAME, EXPORTS_DIR
from models import Answer


@dataclass
class RunSummary:
    total: int
    auto_approved: int
    human_approved: int
    rejected: int
    needs_review: int

    @property
    def auto_pct(self) -> float:
        return (self.auto_approved / self.total * 100) if self.total else 0.0

    @property
    def reviewed_pct(self) -> float:
        return (self.human_approved / self.total * 100) if self.total else 0.0


def summarize_run(answers: List[Answer]) -> RunSummary:
    counts = {"auto_approved": 0, "human_approved": 0, "rejected": 0, "needs_review": 0}
    for a in answers:
        counts[a.status] = counts.get(a.status, 0) + 1
    return RunSummary(
        total=len(answers),
        auto_approved=counts["auto_approved"],
        human_approved=counts["human_approved"],
        rejected=counts["rejected"],
        needs_review=counts["needs_review"],
    )


_STATUS_FILL = {
    "auto_approved": PatternFill("solid", fgColor="D4EDDA"),
    "human_approved": PatternFill("solid", fgColor="CCE5FF"),
    "rejected": PatternFill("solid", fgColor="F8D7DA"),
    "needs_review": PatternFill("solid", fgColor="FFF3CD"),
}


def export_workbook(
    answers: List[Answer], out_dir: Path = EXPORTS_DIR, filename: str | None = None
) -> Path:
    """Write the filled questionnaire to xlsx and return the path."""
    out_dir.mkdir(parents=True, exist_ok=True)
    if filename is None:
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = f"trustloop_{ts}.xlsx"
    out_path = out_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = f"{COMPANY_NAME} Responses"

    headers = [
        "Question ID",
        "Original Question",
        "Status",
        "Answer",
        "Cited Sources",
        "Confidence",
        "Risk Flags",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="343A40")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for a in answers:
        ws.append([
            a.question_id,
            a.question_text,
            a.status,
            a.draft,
            ", ".join(a.evidence) if a.evidence else "—",
            f"{a.confidence:.2f}",
            "; ".join(a.risk_flags) if a.risk_flags else "—",
        ])
        row = ws.max_row
        fill = _STATUS_FILL.get(a.status)
        if fill:
            ws.cell(row=row, column=3).fill = fill
        for col in (2, 4, 5, 7):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {1: 14, 2: 50, 3: 18, 4: 60, 5: 32, 6: 12, 7: 40}
    for col, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    wb.save(out_path)
    return out_path
