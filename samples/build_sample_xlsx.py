"""Generate a sample security questionnaire .xlsx for TrustLoop demos."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

QUESTIONS = [
    # Technical
    "Do you encrypt data at rest?",
    "Do you encrypt data in transit?",
    "Do you support customer-managed encryption keys (BYOK)?",
    "What is your MFA policy?",
    "Do you support SSO via SAML or OIDC?",
    "How do you handle vulnerability management?",
    "Do you perform regular penetration testing?",
    "What is your uptime SLA?",
    "Describe your disaster recovery capabilities.",
    "What logging and monitoring do you provide?",
    # Certification
    "Are you SOC 2 Type II certified?",
    "Are you ISO 27001 certified?",
    "Are you HIPAA certified?",
    "Are you PCI DSS certified?",
    "Are you FedRAMP authorized?",
    # Data Privacy
    "Where is customer data stored?",
    "What is your data retention policy?",
    "Do you have a list of sub-processors?",
    "How do you handle cross-border data transfers?",
    "Can we execute a Data Processing Agreement (DPA)?",
    # Legal
    "Do you guarantee zero security incidents?",
    "Will you indemnify us for data loss?",
    "What are your SLA credits for downtime?",
    "Do you carry cyber liability insurance?",
    # General
    "Do you have a business continuity plan?",
    "How do you train employees on security?",
    "What is your incident response process?",
]


def build_xlsx(out_path: Path | None = None) -> Path:
    if out_path is None:
        out_path = Path(__file__).parent / "sample_questionnaire.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Security Questionnaire"

    headers = ["#", "Question", "Category"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A2D5C")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    categories = {
        0: "Technical", 1: "Technical", 2: "Technical", 3: "Technical",
        4: "Technical", 5: "Technical", 6: "Technical", 7: "Technical",
        8: "Technical", 9: "Technical",
        10: "Certification", 11: "Certification", 12: "Certification",
        13: "Certification", 14: "Certification",
        15: "Data Privacy", 16: "Data Privacy", 17: "Data Privacy",
        18: "Data Privacy", 19: "Data Privacy",
        20: "Legal", 21: "Legal", 22: "Legal", 23: "Legal",
        24: "General", 25: "General", 26: "General",
    }

    for i, q in enumerate(QUESTIONS):
        ws.append([i + 1, q, categories.get(i, "General")])
        row = ws.max_row
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    path = build_xlsx()
    print(f"Generated: {path}")
